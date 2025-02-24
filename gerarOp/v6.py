# -*- coding: utf-8 -*-
from numpy.core.numeric import zeros_like
import openpyxl
import wx
from openpyxl import Workbook
import datetime
import os
import criar_xlsx
from decimal import *
from conexao import *
import varPerda
import varRetrabalho
import numpy
import json

def v6(path):
    pathAp =  path
    tpAp = openpyxl.load_workbook(pathAp, data_only=True)
    tabAp = tpAp.active

    nOP = ''


    cursor = dbConn()
    cursor.execute("SELECT DISTINCT ZKI_KANBAN 'Kanban', RTRIM(CAD.[PN]) 'PN', ZKI_ITEMOP, SB1.B1_PESO, IIF(SB1.B1_PESO > 0, CAST(((1 / SB1.B1_PESO)) AS NUMERIC(16,6)), 0) 'Peca por KG', RTRIM(CAD.[LINHA]) 'LINHA', rtrim(SB1.B1_LOCPAD) 'Armazem' FROM ZKI010 ZKI WITH (NOLOCK) LEFT JOIN (SELECT CASE ISNULL(ZKB.ZKB_KANBAN,'') WHEN '' THEN B1_ZZKNBAN ELSE ZKB.ZKB_KANBAN END [KANBAN], B1_COD [PN], B1_ZZLNPRD [LINHA] FROM SB1010 (NOLOCK) SB1 LEFT JOIN ZKB010 ZKB ON SB1.B1_COD = ZKB.ZKB_PN AND ZKB.D_E_L_E_T_ = '' WHERE SB1.D_E_L_E_T_ = '' AND SB1.B1_ZZKNBAN <> '') [CAD] ON CAD.KANBAN = ZKI.ZKI_KANBAN INNER JOIN SB1010 SB1 ON SB1.B1_COD = CAD.[PN] WHERE ZKI.D_E_L_E_T_ = '' AND ZKI.ZKI_STATUS = 'L';")
    
    vetDados = cursor.fetchall()

##############################################
    # [linha][0] -- Kanban
    # [linha][1] -- Código da Peça
    # [linha][2] -- Código do Item
    # [linha][3] -- Peso da Peça
    # [linha][4] -- PC por KG
    # [linha][5] -- Setor que a peça pertence
    # [linha][6] -- Armazém Padrão
##############################################
    vetDados = numpy.transpose(vetDados)
    resumoAp = {}
    listMovimentacoes = []

    dictItemTry = {
        "t0": '',
        "t1": '',
        "t2": '',
        "t3": '', 
        "t4": '',
        "t6": '',
        "t7": '',
        "t8": '',
        "t9": ''
    }

    # TAMANHO VARIAVEL DE MOVIMENTACOES

    def recuperaPN(kanban):
        resultado = (numpy.where(vetDados[0] == kanban))
        indice = resultado[0][0]
        PartNumber = vetDados[1][indice]
        return indice, PartNumber
    
    def coletaArm(codigo):
        selectCodigo ="SELECT rtrim(B1_COD), B1_LOCPAD FROM SB1010 SB1 (NOLOCK)"
        where = " WHERE UPPER(B1_COD) = '" +codigo+ "' AND D_E_L_E_T_ = ''"
        cursor.execute(selectCodigo + where)
        vetColetaCod = cursor.fetchall()
        armazem = vetColetaCod[0][1]
        return armazem 
    
    def geraItem(codigo):
        for item in dictItemTry.keys():
            if dictItemTry[item] == '':
                dictItemTry[item] = codigo
                opItem = item
                break
    
        return opItem

    def coletaPeso(codigo):
        selectCodigo ="select RTRIM(B1_COD), IIF(SB1.B1_PESO > 0, CAST(((1 / SB1.B1_PESO)) AS NUMERIC(16,6)), 0) 'Peca por KG' from SB1010 SB1 (NOLOCK)"
        where = " WHERE B1_COD = '" +codigo+ "'"
        cursor.execute(selectCodigo + where)
        vetColetaPeso = cursor.fetchall()
        peso = vetColetaPeso[0][1]
        return peso 
    
    def coletaMovimentacao(dados):
        verifica = False

        for i in range(len(listMovimentacoes)):
            componente = listMovimentacoes[i][1]
            tipoMov = listMovimentacoes[i][3]
            quantidade = 0
            origem = listMovimentacoes[i][4]
            destino = listMovimentacoes[i][5]
            auxVerifica = [componente, tipoMov, origem, destino]

            dadosVerifica = [dados[1], dados[3], dados[4], dados[5]]
            # print(dadosVerifica)
            # print(auxVerifica)
            
            if(auxVerifica == dadosVerifica):
                listMovimentacoes[i][2] += dados[2]
                verifica = True
                break

            else:
                verifica = False
                
        if(verifica == False):
            listMovimentacoes.append(dados)
           

    for k in range(12,tabAp.max_row+1):
        verificaTry = str(tabAp['A'+str(k)].value).upper()
        kanban = str(tabAp['B'+str(k)].value).upper()
        maquina = str(tabAp['D'+str(k)].value).upper()
        PartNumber = ''

        verificaKanban = kanban in vetDados[0]

        if(verificaKanban or (verificaTry == "T" and kanban != "TRY OUT" and kanban != "COMPONENTE")) :
            alimentaKanban = {
                        "PN": '',
                        "Item": '',
                        "Setor": '',
                        "totProd": 0, 
                        "pcsBoas": 0,
                        "Retrabalho": 0,
                        "Senbetsu:": 0,
                        "Descarte": 0,
                        "pcsRuins": 0,
                        "PerdaStp_KG": 0,
                        "Perda_KG": 0,
                        "TotalPerda_KG":0,
                        "totKg_Pcs": 0,
                        "pecaPorKG": 0,
                        "classificaPerdas": {}
            }
            
            if(verificaTry == "T" and kanban != "TRY OUT"):     
                PartNumber = kanban

            # APONTAMENTO PADRÃO
            if(maquina != "RETRABALHO" and verificaTry != "R"):   

                def inicializaKanban(kanban, PartNumber):
                    if kanban not in resumoAp:
                        resumoAp[kanban] = alimentaKanban
                        
                        #FAZ BUSCA DO KANBAN NA VARIAVEL VETDADOS E ARMAZENA O INDICE.
                        if(verificaKanban):
                            indice, PartNumber = recuperaPN(kanban)
                            setor = vetDados[5][indice]
                            item = vetDados[2][indice]
                            pecaPorKG = float(vetDados[4][indice])
                        else:
                            setor = str(tabAp['CD'+str(k)].value).upper()
                            item = geraItem(PartNumber)
                            pecaPorKG = coletaPeso(PartNumber)
                            
                        #INICIALIZA SETOR
                        resumoAp[kanban]["Setor"] = setor
                        #INICIALIZA PERDAS
                        classificaDefeitos = varPerda.iniPerdas(setor)
                        resumoAp[kanban]["classificaPerdas"] = classificaDefeitos
                        #INICIALIZA PERDAS
                        classificaRetrabalho = varRetrabalho.iniRetrabalho(setor)
                        
                        resumoAp[kanban]["pecaPorKG"] = pecaPorKG
                        resumoAp[kanban]["classificaRetrabalho"] = classificaRetrabalho

                        resumoAp[kanban]["PN"] = PartNumber
                        resumoAp[kanban]["Item"] = item

                    if(str(tabAp['E'+str(k)].value) != "None"): 
                        #TOTAL PRODUZIDO
                        resumoAp[kanban]["totProd"] += float(tabAp['E'+str(k)].value)

                        # PECAS BOAS
                        if(float(tabAp['AG'+str(k)].value) > 0):  
                            resumoAp[kanban]["pcsBoas"] += float(tabAp['AG'+str(k)].value)
                        
                        #PECAS DESCARTADAS
                        if(int(tabAp['AC'+str(k)].value) != "None"):              
                            resumoAp[kanban]["Descarte"] += int(tabAp['AC'+str(k)].value)
                            
                        #PECAS COM DEFEITO PARA RETRABALHO
                        if(int(tabAp['Q'+str(k)].value) != "None"):              
                            resumoAp[kanban]["Retrabalho"]  += int(tabAp['Q'+str(k)].value)
                        
                        if(int(tabAp['AD'+str(k)].value) != "None"):              
                            resumoAp[kanban]["pcsRuins"]  += int(tabAp['AD'+str(k)].value)
                        
                        
                        clRetrabalho = 0
                        # PEGA DADOS REFERENTE A CLASSIFICAÇÃO DE RETRABALHO
                        for d in range(ord('G'), ord('P')):
                            qtdRetrabalho = tabAp[chr(d)+str(k)].value
                            if(str(qtdRetrabalho) != "None" and qtdRetrabalho > 0):
                                for i, tipoRetrabalho in enumerate(resumoAp[kanban]["classificaRetrabalho"]):
                                    if(i == clRetrabalho):
                                        resumoAp[kanban]["classificaRetrabalho"][tipoRetrabalho] += int(qtdRetrabalho)
                                        break
                            clRetrabalho+=1


                        # PEGA DADOS REFERENTE A CLASSIFICAÇÃO DE DEFEITOS #
                        clDefeitos = 0
                        for d in range(18, 28+1):
                            qtdDef = tabAp.cell(row=k, column=d).value
                            if(str(qtdDef) != "None" and qtdDef > 0):
                                for i, tipoDef in enumerate(resumoAp[kanban]["classificaPerdas"]):
                                    if(i == clDefeitos):
                                        resumoAp[kanban]["classificaPerdas"][tipoDef] += int(qtdDef)
                                        break
                            clDefeitos+=1

                        # TRATAMENTO TOTAL PERDIDO EM KG
                        pecaPorKG = float(resumoAp[kanban]["pecaPorKG"])
                        if(str(tabAp['AE'+str(k)].value) != "None" and tabAp['AE'+str(k)].value > 0):                           

                            # PERDA SETUP EM KG
                            resumoAp[kanban]["PerdaStp_KG"] += float(tabAp['AE'+str(k)].value)
                            resumoAp[kanban]["TotalPerda_KG"] += float(tabAp['AE'+str(k)].value)

                            resumoAp[kanban]["totKg_Pcs"] += float(tabAp['AE'+str(k)].value) * pecaPorKG


                        if(str(tabAp['AF'+str(k)].value) != "None" and tabAp['AF'+str(k)].value > 0): 
                            # PERDA EM KG DO PROCESSO
                            resumoAp[kanban]["Perda_KG"] += float(tabAp['AF'+str(k)].value)
                            resumoAp[kanban]["TotalPerda_KG"] += float(tabAp['AF'+str(k)].value)

                            resumoAp[kanban]["totKg_Pcs"] += float(tabAp['AF'+str(k)].value) * pecaPorKG   
                        #################################
                inicializaKanban(kanban, PartNumber)


            # RETORNO RETRABALHO
            if(maquina == "RETRABALHO" or verificaTry == "R"):
                retBoa = 0
                retRuim = 0

                if(str(tabAp['AG'+str(k)].value) != "None" and tabAp['AG'+str(k)].value > 0):
                    retBoa = tabAp['AG'+str(k)].value
                
                if(str(tabAp['AC'+str(k)].value) != "None" and tabAp['AC'+str(k)].value > 0):
                    retRuim = tabAp['AC'+str(k)].value
                
                #FAZ BUSCA DO KANBAN NA VARIAVEL VETDADOS E ARMAZENA O INDICE.
                if(verificaKanban):
                    indice, PartNumber = recuperaPN(kanban)
                    destino = vetDados[6][indice]
                else:
                    destino = coletaArm(PartNumber)                 

                if(retBoa > 0):
                    coletaMovimentacao([kanban, PartNumber, retBoa, "TRANSFERÊNCIA", "25", destino])

                if(retRuim > 0):
                    coletaMovimentacao([kanban, PartNumber, retRuim, "BAIXA", "25", ""])
                
        ## REGISTRA MOVIMENTAÇÕES TRANSFERENCIA 
        if(str(tabAp['BX'+str(k)].value) != "None"):
            setorDest = ''
            coletaCod = ''
            kanban = ''
            setorOrigem = ''
            indSenbetsu = False

            indice = False
            # print("Teve quantidade preenchida.")
            qtdMovimentacao = float(tabAp['BX'+str(k)].value)

            if(str(tabAp['BV'+str(k)].value) != "None"):
                coletaCod = str(tabAp['BV'+str(k)].value).upper()
            else:
                coletaCod = str(tabAp['B'+str(k)].value).upper()

            # Verifica se o apontamento em questão é um senbetsu

            if(str(tabAp['B'+str(k)].value).upper() == "SENBETSU"):
                indSenbetsu = True

            if coletaCod in vetDados[0]:  
                kanban = coletaCod
                indice, coletaCod = recuperaPN(coletaCod)

            tipoMovimentacao = str(tabAp['BZ'+str(k)].value)


            def verificaSetor(nomeSetor):
                if(nomeSetor == "QUALIDADE"):
                    codSetor = "05"
                elif(nomeSetor == "ENGENHARIA"):
                    codSetor = "03"
                elif(nomeSetor == "RETRABALHO"):
                    codSetor = "25"
                elif(nomeSetor == "SERVICE PARTS"):
                    codSetor = "26"
                elif(nomeSetor == "SENBETSU"):
                    codSetor = "13"
                else:
                    if(indice != False):
                        codSetor = vetDados[6][indice]
                    else:
                        codSetor = coletaArm(coletaCod)
                
                return codSetor
            

            if(str(tabAp['CB'+str(k)].value) != "None"):
                coletaDest = str(tabAp['CB'+str(k)].value).upper()
                setorDest = verificaSetor(coletaDest)
                

            if(indSenbetsu):
                setorOrigem = "13"
            else:   
                coletaOrigem = str(tabAp['CD'+str(k)].value).upper()     
                setorOrigem = verificaSetor(coletaOrigem)

            if(tipoMovimentacao == "TRANSFERÊNCIA" and setorOrigem == setorDest):
                print("O kanban "+kanban+ " com código "+coletaCod + " tem origem igual ao destino")
            else:  
                coletaMovimentacao([kanban, coletaCod, qtdMovimentacao, tipoMovimentacao, setorOrigem, setorDest])
            
            # listMovimentacoes.append([kanban, coletaCod, qtdMovimentacao, tipoMovimentacao, setorOrigem, setorDest])



############# BAIXA DE KANBANS/COMPONENTES #############

    ## PERCORRE A PLANILHA EM BUSCA DE APONTAMENTOS DE BAIXA

    ##atribuição do caminho selecionado pelo usuário para passar os apontamentos do dia#

    dlg = wx.TextEntryDialog(None, 'Qual número da OP?','Dialog')   
    if dlg.ShowModal() == wx.ID_OK: 
        nOP = str(dlg.GetValue()) 
    dlg.Destroy() 

    dlg = wx.TextEntryDialog(None, 'Informe o dia: XX/XX/XXXX','Dialog')   
    if dlg.ShowModal() == wx.ID_OK: 
        data = str(dlg.GetValue()) 
    dlg.Destroy() 
    
    with wx.DirDialog(None, "Selecione o Diretório.", "", wx.DD_DEFAULT_STYLE | wx.DD_DIR_MUST_EXIST) as fileDialog:
        if fileDialog.ShowModal() == wx.ID_CANCEL:
            return     # the user changed their mind"
        pathNew = fileDialog.GetPath()

    dataCorrigida = data.replace("/","-")



    
    # #CRIAR ARQUIVO PARA ARMAZENAR OS APONTAMENTOS#

    criar_xlsx.xlsx(nOP, pathNew)  
    # xlsx = open(pathNew+"\\AP - "+ nOP+".xlsx", 'w') 

    pathValida = pathNew+'\\AP - '+nOP+'.xlsx'
    valida = openpyxl.load_workbook(pathValida)
    tbValida = valida.active
    
    arq = open(pathValida)

    lin = 4
    tbValida['B3'] = "Kanban"
    tbValida['C3'] = "Código"
    tbValida['D3'] = "Setor"
    tbValida['E3'] = "Peças Boas"
    tbValida['F3'] = "Descarte"
    tbValida['G3'] = "Retrabalho"
    tbValida['H3'] = "Total Peças"

    tbValida['I3'] = "Perda Setup KG"
    tbValida['J3'] = "Perda KG"
    tbValida['K3'] = "Total Perda KG"
    tbValida['L3'] = "Total Perda KG/PC"

    for kanban in resumoAp.keys():
        tbValida['B'+str(lin)] = kanban
        tbValida['C'+str(lin)] = resumoAp[kanban]["PN"]
        tbValida['D'+str(lin)] = resumoAp[kanban]["Setor"]

        tbValida['E'+str(lin)] = round(resumoAp[kanban]["pcsBoas"],2)
        tbValida['F'+str(lin)] = round(resumoAp[kanban]["Descarte"],2)
        tbValida['G'+str(lin)] = round(resumoAp[kanban]["Retrabalho"],2)
        tbValida['H'+str(lin)] = round(resumoAp[kanban]["totProd"],2)
        tbValida['I'+str(lin)] = round(resumoAp[kanban]["PerdaStp_KG"],2)
        tbValida['J'+str(lin)] = round(resumoAp[kanban]["Perda_KG"],2)
        tbValida['K'+str(lin)] = round(resumoAp[kanban]["TotalPerda_KG"],2)
        tbValida['L'+str(lin)] = round(resumoAp[kanban]["totKg_Pcs"],2)
        
        lin+=1
        
    arq.close()
    valida.save(pathValida)

    # #CRIAR OP
    
    x = open(pathNew+"\\01 - OP "+ nOP+ " " + dataCorrigida+".txt", 'w')
    for kanban in resumoAp.keys():
        if(resumoAp[kanban]["totProd"] > 0):          
            lin = str(nOP+";"+resumoAp[kanban]["Item"]+";001;"+resumoAp[kanban]["PN"] +";" + str(round(resumoAp[kanban]["totProd"],2))+";"+data +";"+data+"\n")
            x.write(lin)

    x.close()    

    # # #APONTAMENTO DE PRODUÇÃO   

# #CRIAR ARQUIVO PARA ARMAZENAR OS APONTAMENTOS#  

    z = open(pathNew+"\\03 - AP "+nOP+" "+ dataCorrigida+".txt", 'w')

    for kanban in resumoAp.keys():
    # for u in range(0, len(vetKanban)):
        if(resumoAp[kanban]["totProd"] > 0):   
            totalProduzido = round(resumoAp[kanban]["totProd"],2)
            linha = (nOP+resumoAp[kanban]["Item"]+"001;"+ str(round(totalProduzido,2)) + "\n")
            z.write(linha)
        else:
            continue
    
    z.close()



    #APONTAMENTO DE PERDA                           
    cursor.execute("SELECT SB1.B1_ZZKNBAN 'Kanban', rtrim(SG1.G1_COD) 'Código Pai', rtrim(SG1.G1_COMP) 'Cód Componente',  SG1.G1_QUANT 'Quantidade', 0 'Peças Perdidas', 0 'Consumo Total'  FROM  SG1010 SG1  INNER JOIN  SB1010 SB1 ON SG1.G1_COD = SB1.B1_COD AND SB1.D_E_L_E_T_ = ' ' AND SB1.B1_MSBLQL = '2' LEFT JOIN  (SELECT G1_COD,G1_COMP, B1_VM_I, B1_UM FROM SG1010 SG1   INNER JOIN SB1010 AS SB1 ON G1_COMP = B1_COD AND SB1.D_E_L_E_T_ = ''  WHERE SG1.D_E_L_E_T_ = ' ') SG1B ON SG1.G1_COD+SG1.G1_COMP = SG1B.G1_COD+SG1B.G1_COMP  WHERE SG1.D_E_L_E_T_ = '' and SG1B.B1_UM != 'HR'  ORDER BY SB1.B1_ZZKNBAN;")
    
    vetPerda = cursor.fetchall()

    # # #Inicializa looping até o último valor preenchido de perdas
    for kanban in resumoAp.keys():
    # for l in range(0, len(vetTotalPerdaKG)):
        if(resumoAp[kanban]["TotalPerda_KG"] > 0):
            for u in range(0, len(vetPerda)):
                if(resumoAp[kanban]["PN"] == vetPerda[u][1]):
                    vetPerda[u][4] += round(resumoAp[kanban]["totKg_Pcs"],6)
                    vetPerda[u][5] += round(vetPerda[u][3],6) * vetPerda[u][4] 

    # GERAÇÃO DE ARQUIVO DE PERDA ##

    pathArq = pathNew+'\\02 - PERDA '+nOP+'.txt'
    
    arq = open(pathArq,'w')

    for kanban in resumoAp.keys():
        # CLASSIFICAÇÃO E ADIÇÃO DE LINHAS DAS PEÇAS PERDIDAS (Defeitos)
        if(resumoAp[kanban]["pcsRuins"] > 0 or resumoAp[kanban]["TotalPerda_KG"] > 0):
            arq.write("A;"+nOP+resumoAp[kanban]["Item"]+"001\n")
            if(resumoAp[kanban]["pcsRuins"] > 0):
                for defRetrabalho in resumoAp[kanban]["classificaRetrabalho"].keys():
                    if(resumoAp[kanban]["classificaRetrabalho"][defRetrabalho] > 0):
                        arq.write('B;'+resumoAp[kanban]["PN"]+";;"+defRetrabalho+";"+ str(resumoAp[kanban]["classificaRetrabalho"][defRetrabalho]) +';' + resumoAp[kanban]["PN"]+ ';25'+ "\n")
  
                for defMotivo in resumoAp[kanban]["classificaPerdas"].keys():
                    if(resumoAp[kanban]["classificaPerdas"][defMotivo] > 0):
                        arq.write('B;'+resumoAp[kanban]["PN"]+";;" +defMotivo+";"+ str(resumoAp[kanban]["classificaPerdas"][defMotivo]) +';;' +"\n")
                
                
            if(resumoAp[kanban]["TotalPerda_KG"] > 0):    
                    for j in range(0, len(vetPerda)):
                        if(resumoAp[kanban]["PN"] == vetPerda[j][1]):
                            valor = Decimal(vetPerda[j][5])
                            valorDec = float(valor)
                            vnumdec = truncate(valorDec,6)

                            if(valorDec > 0):
                                arq.write('B;'+str(vetPerda[j][2])+';;FH;'+str(vnumdec)+';;'+ "\n")
        
    arq.close()  

    ############ ARQUIVO DE TRANSFERENCIA DE KANBANS/COMPONENTES #############
    # CRIA ARQUIVO DE TRANSFERENCIAS
    
    if(len(listMovimentacoes) > 0):

        y = open(pathNew+"\\04 - TF "+ nOP + " " + dataCorrigida +".txt", 'w')
        w = open(pathNew+"\\05- BAIXA "+ nOP+ " " + dataCorrigida+".txt", 'w')  

        for k in range(0,len(listMovimentacoes)):
            origem = str(listMovimentacoes[k][4])
            destino = str(listMovimentacoes[k][5])
            if(listMovimentacoes[k][3] == "TRANSFERÊNCIA"):
                
                linTransf = str(listMovimentacoes[k][1]+";"+str(listMovimentacoes[k][2])+";" + origem + ";" + destino + ";" + "TF"+ nOP + "\n")
                
                y.write(linTransf)

            ############# ARQUIVO DE BAIXA DE KANBANS/COMPONENTES #############
            if(listMovimentacoes[k][3] == "BAIXA"):
                # CRIA ARQUIVO DE BAIXAS                 
                linBaixa = str(data+";"+listMovimentacoes[k][1]+";"+str(listMovimentacoes[k][2])+";" + origem + "\n")

                w.write(linBaixa)
                #####################################################################

        w.close()
        y.close()  
        

            

        

def truncate(f, n):
    '''Truncates/pads a float f to n decimal places without rounding'''
    s = '{}'.format(f)
    if 'e' in s or 'E' in s:
        return '{0:.{1}f}'.format(f, n)
    i, p, d = s.partition('.')
    return '.'.join([i, (d+'0'*n)[:n]])


