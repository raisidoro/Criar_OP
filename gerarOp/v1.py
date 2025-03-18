# -*- coding: utf-8 -*-
from numpy.core.numeric import zeros_like
import openpyxl as xl
import wx
from openpyxl import Workbook
import datetime
from decimal import *
from conexao import *
import numpy
from pathlib import Path

def v01(arquivo,data):
    cursor   = dbConn()
    cursor.execute("SELECT DISTINCT ZKI_KANBAN 'Kanban', RTRIM(CAD.[PN]) 'PN', ZKI_ITEMOP, SB1.B1_PESO, IIF(SB1.B1_PESO > 0, CAST(((1 / SB1.B1_PESO)) AS NUMERIC(16,6)), 0) 'Peca por KG', RTRIM(CAD.[LINHA]) 'LINHA', rtrim(SB1.B1_LOCPAD) 'Armazem' FROM ZKI010 ZKI WITH (NOLOCK) LEFT JOIN (SELECT CASE ISNULL(ZKB.ZKB_KANBAN,'') WHEN '' THEN B1_ZZKNBAN ELSE ZKB.ZKB_KANBAN END [KANBAN], B1_COD [PN], B1_ZZLNPRD [LINHA] FROM SB1010 (NOLOCK) SB1 LEFT JOIN ZKB010 ZKB WITH (NOLOCK) ON SB1.B1_COD = ZKB.ZKB_PN AND ZKB.D_E_L_E_T_ = '' WHERE SB1.D_E_L_E_T_ = '' AND SB1.B1_ZZKNBAN <> '') [CAD] ON CAD.KANBAN = ZKI.ZKI_KANBAN INNER JOIN SB1010 SB1 WITH (NOLOCK) ON SB1.B1_COD = CAD.[PN] WHERE ZKI.D_E_L_E_T_ = '' AND ZKI.ZKI_STATUS = 'L';")
    nOP      = data[8:10] + data[3:5] + data[0:2]  
    valor    = []
    vetDados = cursor.fetchall()
    vetDados = numpy.transpose(vetDados)

    log = open("C:\TOTVS\log" + nOP + ".txt", "w")
    op = open("C:\TOTVS\op" + nOP + ".txt", "w")

    for path in arquivo:
        pathAp     = path
        tpAp       = xl.load_workbook(pathAp, data_only=True)
        tabAp      = tpAp.active
        inicColuna = 1
        kanbanF    = []
        caminho    = []
        i = 0
        j = 5
        
        wb1 = xl.load_workbook(pathAp, data_only = True)

        #aba resumo não foi encontrada ou a aba resumo não pode ser acessada
        try:
            ws1 = wb1['RESUMO']
        except KeyError:
            log.write(f"[{datetime.datetime.now()}] Erro: Aba 'RESUMO' não foi encontrada no arquivo\n")
        except Exception as e:
            log.write(f"[{datetime.datetime.now()}] Erro: Falha ao acessar a aba 'RESUMO' no arquivo, Erro: {str(e)}\n")


        while str(ws1.cell(4,inicColuna).value) != 'KANBAN':
            inicColuna = inicColuna + 1

        i = inicColuna 

        #se a data não estiver na planilha
        i = 1
        nEncontrada = False

        while ws1.cell(3,i).value == None:
            i = i + 1

        while ws1.cell(3,i).value != None: #executa até encontrar a primeira coluna vazia
            if ws1.cell(3,i).value != None:
                if  data[0:5] in str(format(ws1.cell(3,i).value, "%d/%m")):
                    nEncontrada = True
                    break
            i = i + 1
                
        if not nEncontrada:
            log.write(f"[{datetime.datetime.now()}] Erro: Não foi possível encontrar a data {data} na planilha {path}\n")
            print("Data não encontrada na planilha")

        #se a planilha possuir um kanban simultâneo com uma / no meio (ex: O-208/9)
        while ws1.cell(j, inicColuna).value is not None or ws1.cell(j, inicColuna).coordinate in ws1.merged_cells:
            if ws1.cell(j, i).value not in ('', 0, None):
                if 'RESUMO' in str(ws1.cell(j, i).value):
                    j += 4
                    continue
                if '/' in str(ws1.cell(j, inicColuna).value):
                    caminho = str(path)
                    separaKanban = str(ws1.cell(j, inicColuna).value).split('/')

                    primeiroKanban = separaKanban[0].strip()
                    certo = primeiroKanban + ' , ' + str(ws1.cell(j,i).value)
                    valor.append((certo, caminho))

                    prefixo = primeiroKanban[:-len(separaKanban[1].strip())]
                    segundoNumero = separaKanban[1].strip()
                    segundoKanban = prefixo + segundoNumero
                    certo = segundoKanban + ' , ' + str(ws1.cell(j,i).value)
                    valor.append((certo, caminho))

                else:
                    caminho = str(path)
                    certo = (str(ws1.cell(j,inicColuna).value) + ' , ' + str(ws1.cell(j,i).value))
                    valor.append((certo, caminho))   
            j += 1
        
    for valores, caminho in valor: 

        resultado = (numpy.where(vetDados[0] == valores[0:5]))
        
        if len(resultado[0]) == 1:
            indice = resultado[0][0]
            PartNumber = vetDados[1][indice]
            indice = vetDados[2][indice]
        else:
            log.write(f"[{datetime.datetime.now()}] Erro: " + valores[0:5] + " não foi encontrado! " + caminho + "\n")

        try:
            float(valores[8:])
        except ValueError:
            log.write(f"[{datetime.datetime.now()}] Erro: A quantidade informado no produto " + valores[:5] + " não é um número - " + caminho + "\n")               
        else:
            op.write(nOP + ";" + indice + ";001;" + PartNumber + ";" + valores[8:] + ";" + data + ";" + data + ";F \n")

    op.close()

    log.close()
    return 